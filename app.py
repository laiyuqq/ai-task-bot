# -*- coding: utf-8 -*-
"""
AI任务中台 - 推广支持Bot
功能：DeepSeek大模型问答 + 知识库管理(含文档上传) + 需求收集 + 访问统计 + 数据分析 + 管理后台
存储：优先PostgreSQL，未配置时降级为JSON文件
"""
import os
import json
import time
import re
import uuid
import requests as http_requests
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 数据库模块
try:
    import db as dbmod
    DB_ENABLED = dbmod.DB_ENABLED
except Exception:
    DB_ENABLED = False
    dbmod = None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get("DATA_DIR", os.path.join(BASE_DIR, "data"))
REQUESTS_FILE = os.path.join(DATA_DIR, "requests.json")
KB_FILE = os.path.join(BASE_DIR, "knowledge_base.json")
DOCS_DIR = os.path.join(DATA_DIR, "docs")
VISITS_FILE = os.path.join(DATA_DIR, "visits.json")
CHAT_LOG_FILE = os.path.join(DATA_DIR, "chat_logs.json")
DOCS_INDEX_FILE = os.path.join(DATA_DIR, "docs_index.json")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(DOCS_DIR, exist_ok=True)

app = Flask(__name__)

# DeepSeek配置
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")
DEEPSEEK_API_URL = "https://api.deepseek.com/v1/chat/completions"
DEEPSEEK_MODEL = os.environ.get("DEEPSEEK_MODEL", "deepseek-chat")

# ============================================================
# 通用 JSON 存储（JSON模式用）
# ============================================================
def _load_json(path, default):
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return default
    return default

def _save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ============================================================
# 数据存储层 —— 双模式：数据库 / JSON
# ============================================================

# --- 知识库 ---
def load_kb():
    if DB_ENABLED:
        return dbmod.db_load_kb()
    return _load_json(KB_FILE, {"meta": {}, "faq": []})

def reload_kb():
    pass

def _save_kb_json(kb_data):
    _save_json(KB_FILE, kb_data)

# --- 文档 ---
def load_docs():
    if DB_ENABLED:
        return dbmod.db_load_docs()
    return _load_json(DOCS_INDEX_FILE, [])

def save_docs_json(data):
    _save_json(DOCS_INDEX_FILE, data)

def get_all_doc_text():
    docs = load_docs()
    parts = []
    for d in docs:
        if d.get("text"):
            parts.append(f"【{d['name']}】\n{d['text']}")
    return "\n\n".join(parts)

# --- 访问记录 ---
def record_visit():
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown")
    if "," in ip:
        ip = ip.split(",")[0].strip()
    path = request.path
    ua = request.headers.get("User-Agent", "")[:200]
    if DB_ENABLED:
        dbmod.db_record_visit(ip, path, ua)
        return
    visits = _load_json(VISITS_FILE, [])
    visits.append({
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "date": datetime.now().strftime("%Y-%m-%d"),
        "ip": ip, "path": path, "ua": ua
    })
    if len(visits) > 5000:
        visits = visits[-5000:]
    _save_json(VISITS_FILE, visits)

def load_visits():
    if DB_ENABLED:
        return dbmod.db_load_visits()
    return _load_json(VISITS_FILE, [])

# --- 问答日志 ---
def record_chat_log(user_msg, bot_answer, matched, source, duration_ms, error=None):
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown")
    if "," in ip:
        ip = ip.split(",")[0].strip()
    if DB_ENABLED:
        dbmod.db_record_chat_log(user_msg, bot_answer, matched, source, duration_ms, ip, error)
        return
    logs = _load_json(CHAT_LOG_FILE, [])
    logs.append({
        "id": uuid.uuid4().hex[:12],
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "date": datetime.now().strftime("%Y-%m-%d"),
        "ip": ip, "user": user_msg, "bot": bot_answer,
        "matched": matched, "source": source,
        "duration_ms": duration_ms, "error": error
    })
    if len(logs) > 5000:
        logs = logs[-5000:]
    _save_json(CHAT_LOG_FILE, logs)

def load_chat_logs():
    if DB_ENABLED:
        return dbmod.db_load_chat_logs()
    return _load_json(CHAT_LOG_FILE, [])

def clear_chat_logs():
    if DB_ENABLED:
        dbmod.db_clear_chat_logs()
        return
    _save_json(CHAT_LOG_FILE, [])

# --- 需求记录 ---
def load_requests():
    if DB_ENABLED:
        return dbmod.db_load_requests()
    return _load_json(REQUESTS_FILE, [])

def add_request(org, contact, rtype, desc, extra=""):
    if DB_ENABLED:
        return dbmod.db_add_request(org, contact, rtype, desc, extra)
    reqs = _load_json(REQUESTS_FILE, [])
    record = {
        "id": str(int(time.time() * 1000)) + uuid.uuid4().hex[:4],
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "org": org, "contact": contact, "type": rtype,
        "description": desc, "status": "待处理", "extra": extra
    }
    reqs.append(record)
    _save_json(REQUESTS_FILE, reqs)
    return record

def update_request_status(rid, status):
    if DB_ENABLED:
        return dbmod.db_update_request_status(rid, status)
    reqs = load_requests()
    for r in reqs:
        if r["id"] == rid:
            r["status"] = status
            _save_json(REQUESTS_FILE, reqs)
            return True
    return False

def delete_request(rid):
    if DB_ENABLED:
        return dbmod.db_delete_request(rid)
    reqs = load_requests()
    new_reqs = [r for r in reqs if r["id"] != rid]
    if len(new_reqs) == len(reqs):
        return False
    _save_json(REQUESTS_FILE, new_reqs)
    return True

# --- 知识库管理 ---
def add_kb_item(item):
    if DB_ENABLED:
        dbmod.db_add_kb_item(item)
        return
    kb = load_kb()
    kb["faq"].append(item)
    _save_json(KB_FILE, kb)

def update_kb_item(kid, data):
    if DB_ENABLED:
        return dbmod.db_update_kb_item(kid, data)
    kb = load_kb()
    for item in kb["faq"]:
        if item["id"] == kid:
            if "question" in data: item["question"] = data["question"].strip()
            if "answer" in data: item["answer"] = data["answer"].strip()
            if "keywords" in data:
                kws = data["keywords"]
                item["keywords"] = [k.strip() for k in re.split(r"[，,、\n]", kws) if k.strip()] if isinstance(kws, str) else kws
            if "category" in data: item["category"] = data["category"].strip()
            _save_json(KB_FILE, kb)
            return True
    return False

def delete_kb_item(kid):
    if DB_ENABLED:
        return dbmod.db_delete_kb_item(kid)
    kb = load_kb()
    new_faq = [f for f in kb["faq"] if f["id"] != kid]
    if len(new_faq) == len(kb["faq"]):
        return False
    kb["faq"] = new_faq
    _save_json(KB_FILE, kb)
    return True

# --- 文档管理 ---
def add_doc(doc_id, name, file_path, size, text):
    if DB_ENABLED:
        dbmod.db_add_doc(doc_id, name, file_path, size, text)
        return
    docs = load_docs()
    docs.append({
        "id": doc_id, "name": filename, "path": file_path, "size": size,
        "text": text, "upload_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })
    save_docs_json(docs)

def delete_doc(doc_id):
    if DB_ENABLED:
        ok, file_path = dbmod.db_delete_doc(doc_id)
        if ok and file_path:
            try: os.remove(os.path.join(DOCS_DIR, file_path))
            except: pass
        return ok
    docs = load_docs()
    new_docs = [d for d in docs if d["id"] != doc_id]
    if len(new_docs) == len(docs):
        return False
    for d in docs:
        if d["id"] == doc_id:
            try: os.remove(os.path.join(DOCS_DIR, d.get("path", "")))
            except: pass
    save_docs_json(new_docs)
    return True

# ============================================================
# 关键词匹配引擎（作为兜底）
# ============================================================
def normalize(text):
    text = text.lower().strip()
    text = re.sub(r"[\s，。、！？,.!?~～\-_]", "", text)
    return text

def score_match(user_input, faq_item):
    score = 0
    norm_input = normalize(user_input)
    if not norm_input:
        return 0
    for kw in faq_item.get("keywords", []):
        kw_n = normalize(kw)
        if not kw_n:
            continue
        if kw_n in norm_input:
            score += 1 + len(kw_n) * 0.3
    q_n = normalize(faq_item.get("question", ""))
    if q_n and q_n in norm_input:
        score += 5
    if q_n:
        overlap = sum(1 for ch in q_n if ch in norm_input)
        score += (overlap / max(len(q_n), 1)) * 2
    return score

def find_answer_keyword(user_input):
    kb = load_kb()
    best_score = 0
    best_item = None
    for item in kb["faq"]:
        s = score_match(user_input, item)
        if s > best_score:
            best_score = s
            best_item = item
    if best_score < 1.5 or best_item is None:
        return None, 0
    return best_item, best_score

def build_kb_context():
    kb = load_kb()
    parts = []
    for item in kb["faq"]:
        parts.append(f"Q: {item['question']}\nA: {item['answer']}")
    doc_text = get_all_doc_text()
    if doc_text:
        parts.append(f"以下是补充资料：\n{doc_text}")
    return "\n\n".join(parts)

# ============================================================
# DeepSeek 大模型问答
# ============================================================
SYSTEM_PROMPT = """你是「AI任务中台推广支持Bot」，负责为各机构使用者解答关于"AI任务中台"的问题。

你的角色和语气：
- 热情主动的客服助手，让使用者感受到被重视和欢迎
- 回答时先肯定使用者的问题很有价值，再给出清晰、有条理的解答
- 语气亲切自然，像一个靠谱的同事在耐心帮忙，不卑不亢
- 保持理性专业，基于知识库内容回答，不编造信息，不过度承诺
- 回答用中文，适当使用分点、加粗让信息更清晰
- 如果用户的问题在知识库中找不到答案，坦诚告知并热情引导用户点击「我要提需求」按钮反馈，让使用者觉得提需求是被欢迎的
- 不要提及你是AI模型或你的技术实现细节
- 项目经理是赖雨晴，涉及联系/开通试用等问题时提及她
- 回答开头可以适当用"好的！""没问题！"等热情的回应，但不要每句都这样，自然就好

知识库内容如下：
{kb_context}
"""

def call_deepseek(user_msg, history=None):
    if not DEEPSEEK_API_KEY:
        return None, "未配置DEEPSEEK_API_KEY"
    kb_context = build_kb_context()
    system_content = SYSTEM_PROMPT.replace("{kb_context}", kb_context)
    messages = [{"role": "system", "content": system_content}]
    if history:
        for h in history[-10:]:
            messages.append({"role": h.get("role", "user"), "content": h.get("content", "")})
    messages.append({"role": "user", "content": user_msg})
    try:
        resp = http_requests.post(
            DEEPSEEK_API_URL,
            headers={"Authorization": f"Bearer {DEEPSEEK_API_KEY}", "Content-Type": "application/json"},
            json={"model": DEEPSEEK_MODEL, "messages": messages, "max_tokens": 1024, "temperature": 0.3, "stream": False},
            timeout=30
        )
        if resp.status_code == 200:
            data = resp.json()
            answer = data["choices"][0]["message"]["content"]
            return answer, None
        else:
            return None, f"API返回{resp.status_code}: {resp.text[:200]}"
    except Exception as e:
        return None, str(e)

# ============================================================
# 中间件 - 访问统计
# ============================================================
@app.before_request
def _track_visit():
    if not request.path.startswith("/static"):
        try:
            record_visit()
        except Exception:
            pass  # 访问记录失败不应阻断请求

# ============================================================
# 路由 - 页面
# ============================================================
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/admin")
def admin():
    return render_template("admin.html")

# ============================================================
# API - 问答
# ============================================================
@app.route("/api/chat", methods=["POST"])
def api_chat():
    data = request.get_json(silent=True) or {}
    user_msg = data.get("message", "").strip()
    history = data.get("history", [])
    if not user_msg:
        return jsonify({"ok": False, "error": "消息不能为空"}), 400

    start_time = time.time()

    if DEEPSEEK_API_KEY:
        answer, err = call_deepseek(user_msg, history)
        duration_ms = int((time.time() - start_time) * 1000)
        if answer:
            try: record_chat_log(user_msg, answer, True, "deepseek", duration_ms)
            except: pass
            return jsonify({"ok": True, "matched": True, "answer": answer, "source": "deepseek", "duration_ms": duration_ms})
        item, score = find_answer_keyword(user_msg)
        if item:
            duration_ms = int((time.time() - start_time) * 1000)
            try: record_chat_log(user_msg, item["answer"], True, "keyword_fallback", duration_ms, error=err)
            except: pass
            return jsonify({"ok": True, "matched": True, "answer": item["answer"], "source": "keyword", "category": item.get("category", ""), "duration_ms": duration_ms})
        duration_ms = int((time.time() - start_time) * 1000)
        fallback_msg = "抱歉，这个问题我暂时还没有现成的答案。\n\n您可以通过下方的「我要提需求」按钮，把问题或需求留给我们，项目经理赖雨晴会尽快跟进。"
        try: record_chat_log(user_msg, fallback_msg, False, "fallback", duration_ms, error=err)
        except: pass
        return jsonify({"ok": True, "matched": False, "answer": fallback_msg, "suggest_feedback": True})

    item, score = find_answer_keyword(user_msg)
    duration_ms = int((time.time() - start_time) * 1000)
    if item:
        try: record_chat_log(user_msg, item["answer"], True, "keyword", duration_ms)
        except: pass
        return jsonify({"ok": True, "matched": True, "answer": item["answer"], "category": item.get("category", ""), "source": "keyword", "duration_ms": duration_ms})
    fallback_msg = "抱歉，这个问题我暂时还没有现成的答案。\n\n您可以通过下方的「我要提需求」按钮，把问题或需求留给我们，项目经理赖雨晴会尽快跟进。常见问题也可以点击下方快捷提问试试。"
    try: record_chat_log(user_msg, fallback_msg, False, "fallback", duration_ms)
    except: pass
    return jsonify({"ok": True, "matched": False, "answer": fallback_msg, "suggest_feedback": True})

@app.route("/api/quick_questions", methods=["GET"])
def api_quick_questions():
    kb = load_kb()
    items = [{"id": f["id"], "question": f["question"]} for f in kb["faq"]]
    return jsonify({"ok": True, "questions": items})

# ============================================================
# API - 需求收集
# ============================================================
@app.route("/api/feedback", methods=["POST"])
def api_feedback():
    data = request.get_json(silent=True) or {}
    org = data.get("org", "").strip()
    contact = data.get("contact", "").strip()
    rtype = data.get("type", "咨询").strip()
    desc = data.get("description", "").strip()
    extra = data.get("extra", "").strip()
    if not desc:
        return jsonify({"ok": False, "error": "问题描述不能为空"}), 400
    record = add_request(org, contact, rtype, desc, extra)
    return jsonify({"ok": True, "message": "已收到您的反馈，项目经理赖雨晴会尽快跟进，感谢支持！", "record_id": record["id"]})

# ============================================================
# API - 管理后台（需求）
# ============================================================
@app.route("/api/requests", methods=["GET"])
def api_requests():
    reqs = load_requests()
    org = request.args.get("org", "").strip()
    rtype = request.args.get("type", "").strip()
    status = request.args.get("status", "").strip()
    result = reqs
    if org: result = [r for r in result if org in r.get("org", "")]
    if rtype: result = [r for r in result if r.get("type") == rtype]
    if status: result = [r for r in result if r.get("status") == status]
    return jsonify({"ok": True, "total": len(result), "data": result})

@app.route("/api/requests/<rid>/status", methods=["POST"])
def api_update_status(rid):
    data = request.get_json(silent=True) or {}
    new_status = data.get("status", "").strip()
    if update_request_status(rid, new_status):
        return jsonify({"ok": True, "message": "状态已更新"})
    return jsonify({"ok": False, "error": "未找到记录"}), 404

@app.route("/api/requests/<rid>", methods=["DELETE"])
def api_delete_request(rid):
    if delete_request(rid):
        return jsonify({"ok": True, "message": "已删除"})
    return jsonify({"ok": False, "error": "未找到记录"}), 404

@app.route("/api/stats", methods=["GET"])
def api_stats():
    reqs = load_requests()
    by_type, by_status, by_org = {}, {}, {}
    for r in reqs:
        t = r.get("type", "未分类"); s = r.get("status", "待处理"); o = r.get("org", "未填写") or "未填写"
        by_type[t] = by_type.get(t, 0) + 1
        by_status[s] = by_status.get(s, 0) + 1
        by_org[o] = by_org.get(o, 0) + 1
    return jsonify({"ok": True, "total": len(reqs), "by_type": by_type, "by_status": by_status, "by_org": by_org})

@app.route("/api/export_excel", methods=["GET"])
def api_export_excel():
    reqs = load_requests()
    wb = Workbook(); ws = wb.active; ws.title = "需求汇总"
    headers = ["序号", "提交时间", "机构名称", "联系人", "类型", "问题描述", "补充说明", "状态"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = thin_border
    for i, r in enumerate(reqs, 1):
        values = [i, r.get("time",""), r.get("org",""), r.get("contact",""), r.get("type",""), r.get("description",""), r.get("extra",""), r.get("status","")]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=i+1, column=col, value=v if v else "")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border
    widths = [6, 20, 18, 14, 10, 50, 30, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+col)].width = w
    ws.freeze_panes = "A2"
    out_path = os.path.join(DATA_DIR, "需求汇总.xlsx")
    wb.save(out_path)
    filename = f"AI任务中台_需求汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(out_path, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ============================================================
# API - 知识库管理（FAQ）
# ============================================================
@app.route("/api/kb", methods=["GET"])
def api_kb_list():
    kb = load_kb()
    return jsonify({"ok": True, "meta": kb.get("meta", {}), "faq": kb["faq"]})

@app.route("/api/kb", methods=["POST"])
def api_kb_add():
    data = request.get_json(silent=True) or {}
    question = data.get("question", "").strip()
    answer = data.get("answer", "").strip()
    keywords_raw = data.get("keywords", "").strip()
    category = data.get("category", "其他").strip()
    if not question or not answer:
        return jsonify({"ok": False, "error": "问题和答案不能为空"}), 400
    keywords = [k.strip() for k in re.split(r"[，,、\n]", keywords_raw) if k.strip()]
    for ch in question:
        if ch not in keywords and ch not in "？?的了吗呢啊是有什么":
            keywords.append(ch)
    new_id = "custom_" + uuid.uuid4().hex[:8]
    item = {"id": new_id, "category": category, "question": question, "keywords": keywords, "answer": answer}
    add_kb_item(item)
    return jsonify({"ok": True, "message": "已新增知识库条目", "id": new_id})

@app.route("/api/kb/<kid>", methods=["PUT"])
def api_kb_update(kid):
    data = request.get_json(silent=True) or {}
    if update_kb_item(kid, data):
        return jsonify({"ok": True, "message": "已修改"})
    return jsonify({"ok": False, "error": "未找到条目"}), 404

@app.route("/api/kb/<kid>", methods=["DELETE"])
def api_kb_delete(kid):
    if delete_kb_item(kid):
        return jsonify({"ok": True, "message": "已删除"})
    return jsonify({"ok": False, "error": "未找到条目"}), 404

# ============================================================
# API - 知识库文档上传
# ============================================================
@app.route("/api/docs", methods=["GET"])
def api_docs_list():
    docs = load_docs()
    result = [{"id": d["id"], "name": d["name"], "size": d.get("size", 0),
               "upload_time": d.get("upload_time", ""), "preview": d.get("text", "")[:200]} for d in docs]
    return jsonify({"ok": True, "total": len(result), "data": result})

@app.route("/api/docs/upload", methods=["POST"])
def api_docs_upload():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "请选择文件"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"ok": False, "error": "请选择文件"}), 400

    filename = f.filename
    ext = os.path.splitext(filename)[1].lower()
    safe_name = f"doc_{uuid.uuid4().hex[:8]}{ext}"
    filepath = os.path.join(DOCS_DIR, safe_name)
    f.save(filepath)
    size = os.path.getsize(filepath)

    text = ""
    try:
        if ext in (".txt", ".md"):
            with open(filepath, "r", encoding="utf-8") as fp:
                text = fp.read()
        elif ext == ".pdf":
            try:
                import pdfplumber
                with pdfplumber.open(filepath) as pdf:
                    text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            except ImportError:
                import subprocess
                result = subprocess.run(["pdftotext", filepath, "-"], capture_output=True, text=True, timeout=30)
                text = result.stdout
        elif ext in (".docx",):
            from docx import Document
            doc = Document(filepath)
            text = "\n".join(para.text for para in doc.paragraphs)
        else:
            text = f"[不支持的文件格式: {ext}]"
    except Exception as e:
        text = f"[文件解析失败: {str(e)}]"

    if len(text) > 50000:
        text = text[:50000] + "\n...（内容过长已截断）"

    doc_id = uuid.uuid4().hex[:12]
    if DB_ENABLED:
        dbmod.db_add_doc(doc_id, filename, safe_name, size, text)
    else:
        docs = load_docs()
        docs.append({"id": doc_id, "name": filename, "path": safe_name, "size": size,
                      "text": text, "upload_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        save_docs_json(docs)

    return jsonify({"ok": True, "message": f"文档「{filename}」上传成功", "id": doc_id, "text_length": len(text)})

@app.route("/api/docs/<doc_id>", methods=["DELETE"])
def api_docs_delete(doc_id):
    if delete_doc(doc_id):
        return jsonify({"ok": True, "message": "已删除"})
    return jsonify({"ok": False, "error": "未找到文档"}), 404

# ============================================================
# API - 访问统计 & 问答日志
# ============================================================
@app.route("/api/analytics/overview", methods=["GET"])
def api_analytics_overview():
    days = request.args.get("days", "7")
    try: days = int(days)
    except: days = 7

    visits = load_visits()
    logs = load_chat_logs()
    today = datetime.now().strftime("%Y-%m-%d")
    start_date = (datetime.now() - timedelta(days=days-1)).strftime("%Y-%m-%d")

    daily_visits = {}
    daily_chats = {}
    for v in visits:
        d = v.get("date", "")
        if d >= start_date:
            daily_visits[d] = daily_visits.get(d, 0) + 1
    for l in logs:
        d = l.get("date", "")
        if d >= start_date:
            daily_chats[d] = daily_chats.get(d, 0) + 1

    date_list = [(datetime.now() - timedelta(days=days-1-i)).strftime("%Y-%m-%d") for i in range(days)]
    daily_visits_arr = [{"date": d, "count": daily_visits.get(d, 0)} for d in date_list]
    daily_chats_arr = [{"date": d, "count": daily_chats.get(d, 0)} for d in date_list]

    today_visits = sum(1 for v in visits if v.get("date") == today)
    today_chats = sum(1 for l in logs if l.get("date") == today)

    return jsonify({"ok": True, "today_visits": today_visits, "today_chats": today_chats,
                    "total_visits": len(visits), "total_chats": len(logs),
                    "daily_visits": daily_visits_arr, "daily_chats": daily_chats_arr})

@app.route("/api/analytics/chat_logs", methods=["GET"])
def api_analytics_chat_logs():
    logs = load_chat_logs()
    logs = list(reversed(logs))
    page = request.args.get("page", "1")
    try: page = int(page)
    except: page = 1
    per_page = 50
    total = len(logs)
    start = (page - 1) * per_page
    end = start + per_page
    return jsonify({"ok": True, "total": total, "page": page, "per_page": per_page, "data": logs[start:end]})

@app.route("/api/analytics/chat_logs", methods=["DELETE"])
def api_analytics_chat_logs_clear():
    clear_chat_logs()
    return jsonify({"ok": True, "message": "已清空"})

# ============================================================
# API - 数据分析（深度洞察）
# ============================================================
@app.route("/api/analytics/insights", methods=["GET"])
def api_analytics_insights():
    reqs = load_requests()
    logs = load_chat_logs()

    # 机构需求排行
    org_count = {}; org_type = {}
    for r in reqs:
        org = r.get("org", "未填写") or "未填写"
        org_count[org] = org_count.get(org, 0) + 1
        if org not in org_type: org_type[org] = {}
        t = r.get("type", "未分类")
        org_type[org][t] = org_type[org].get(t, 0) + 1
    top_orgs = sorted([{"org": k, "count": v, "types": org_type[k]} for k, v in org_count.items()], key=lambda x: x["count"], reverse=True)[:20]

    # 热门问题
    def norm_q(q):
        q = q.lower().strip()
        return re.sub(r"[\s，。、！？,.!?~～\-_？吗呢啊的吧]", "", q)
    q_count = {}; q_examples = {}
    for l in logs:
        u = l.get("user", "").strip()
        if not u: continue
        nk = norm_q(u)
        if nk not in q_count: q_count[nk] = 0; q_examples[nk] = u
        q_count[nk] += 1
    top_questions = sorted([{"question": q_examples[k], "count": v} for k, v in q_count.items()], key=lambda x: x["count"], reverse=True)[:30]

    # 类型分布
    type_dist = {}
    for r in reqs:
        t = r.get("type", "未分类"); type_dist[t] = type_dist.get(t, 0) + 1

    # 来源分布
    source_dist = {}
    source_labels = {"deepseek": "AI大模型", "keyword": "关键词匹配", "keyword_fallback": "关键词(降级)", "fallback": "未匹配"}
    for l in logs:
        s = source_labels.get(l.get("source", ""), l.get("source", "未知"))
        source_dist[s] = source_dist.get(s, 0) + 1

    # 知识盲区
    gaps = []
    for l in logs:
        if not l.get("matched", False):
            gaps.append({"question": l.get("user", ""), "time": l.get("time", ""), "source": l.get("source", ""), "error": l.get("error", "")})
    gaps = list(reversed(gaps))[:50]

    # 响应时间
    durations = [l.get("duration_ms", 0) for l in logs if l.get("duration_ms", 0) > 0]
    if durations:
        ds = sorted(durations); n = len(ds)
        response_time = {"avg": round(sum(ds)/n), "min": ds[0], "max": ds[-1], "p50": ds[n//2], "p95": ds[int(n*0.95)] if n > 1 else ds[0], "count": n}
    else:
        response_time = {"avg": 0, "min": 0, "max": 0, "p50": 0, "p95": 0, "count": 0}

    # 活跃时段
    hour_dist = {}
    for l in logs:
        t = l.get("time", "")
        if len(t) >= 13:
            h = t[11:13]; hour_dist[h] = hour_dist.get(h, 0) + 1
    hourly = [{"hour": f"{h:02d}:00", "count": hour_dist.get(f"{h:02d}", 0)} for h in range(24)]

    return jsonify({"ok": True, "top_orgs": top_orgs, "top_questions": top_questions,
                    "type_distribution": type_dist, "source_distribution": source_dist,
                    "knowledge_gaps": gaps, "gaps_count": len(gaps), "response_time": response_time,
                    "total_requests": len(reqs), "total_chats": len(logs), "hourly_distribution": hourly})


if __name__ == "__main__":
    # 初始化数据库（如配置了）
    if DB_ENABLED:
        try:
            dbmod.init_db()
            # 首次运行：从JSON导入初始知识库到数据库
            if len(load_kb()["faq"]) == 0 and os.path.exists(KB_FILE):
                print("[DB] 从JSON导入初始知识库...")
                kb_data = _load_json(KB_FILE, {"meta": {}, "faq": []})
                for item in kb_data["faq"]:
                    add_kb_item(item)
                print(f"[DB] 已导入 {len(kb_data['faq'])} 条FAQ")
            # 写入示例需求（首次）
            if len(load_requests()) == 0:
                add_request("示例机构A", "张三", "咨询", "想了解AI质检是怎么判断回访有效性的，准确率怎么样？", "")
                add_request("示例机构B", "李四", "需求", "希望话术能支持按产品线区分，不同产品话术不同。", "")
                add_request("示例机构A", "张三", "bug", "总览页面加载比较慢，有时候要等好几秒。", "")
        except Exception as e:
            print(f"[DB] 数据库初始化失败: {e}")
            print("[DB] 降级为JSON文件模式")
            DB_ENABLED = False
            dbmod = None
    else:
        # JSON模式：写入示例数据
        if not os.path.exists(REQUESTS_FILE) or len(load_requests()) == 0:
            add_request("示例机构A", "张三", "咨询", "想了解AI质检是怎么判断回访有效性的，准确率怎么样？", "")
            add_request("示例机构B", "李四", "需求", "希望话术能支持按产品线区分，不同产品话术不同。", "")
            add_request("示例机构A", "张三", "bug", "总览页面加载比较慢，有时候要等好几秒。", "")

    print("=" * 50)
    print("AI任务中台 推广支持Bot 已启动")
    print(f"  存储模式: {'PostgreSQL数据库' if DB_ENABLED else 'JSON文件'}")
    if DB_ENABLED:
        print(f"  数据库: {dbmod.DB_HOST}:{dbmod.DB_PORT}/{dbmod.DB_NAME}")
    print(f"  DeepSeek: {'已配置' if DEEPSEEK_API_KEY else '未配置（纯关键词模式）'}")
    print("  使用者入口: /")
    print("  管理后台:   /admin")
    print("=" * 50)

    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
